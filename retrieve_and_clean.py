from collections import namedtuple
import json
import os.path

from openpyxl import load_workbook
import requests

DATA_DIR = 'source_files/electorate-profiles-2016'
JS_FILE = 'src/states.js'
E_VOTES = {
    'Alabama': 9, 'Alaska': 3, 'Arizona': 11, 'Arkansas': 6, 'California': 55,
    'Colorado': 9, 'Connecticut': 7, 'Delaware': 3, 'District of Columbia': 3,
    'Florida': 29, 'Georgia': 16, 'Hawaii': 4, 'Idaho': 4, 'Illinois': 20,
    'Indiana': 11, 'Iowa': 6, 'Kansas': 6, 'Kentucky': 8, 'Louisiana': 8,
    'Maine': 4, 'Maryland': 10, 'Massachusetts': 11, 'Michigan': 16,
    'Minnesota': 10, 'Mississippi': 6, 'Missouri': 10, 'Montana': 3,
    'Nebraska': 5, 'Nevada': 6, 'New Hampshire': 4, 'New Jersey': 14,
    'New Mexico': 5, 'New York': 29, 'North Carolina': 15, 'North Dakota': 3,
    'Ohio': 18, 'Oklahoma': 7, 'Oregon': 7, 'Pennsylvania': 20,
    'Rhode Island': 4, 'South Carolina': 9, 'South Dakota': 3,
    'Tennessee': 11, 'Texas': 38, 'Utah': 6, 'Vermont': 3, 'Virginia': 13,
    'Washington': 12, 'West Virginia': 5, 'Wisconsin': 10, 'Wyoming': 3,
}

VOTING_AGE_POP = 'B3'
EIGHTEEN_TO_TWENTY_NINE = 'B5'
THIRTY_TO_FOURTY_FOUR = 'B6'
FORTY_FIVE_TO_SIXTY_FOUR = 'B7'
SIXTY_FIVE_AND_OVER = 'B8'
MALE = 'B10'
FEMALE = 'B10'
WHITE = 'B13'
BLACK = 'B14'
NATIVE_AMERICAN = 'B15'
ASIAN = 'B16'
PACIFIC_ISLANDER = 'B17'
OTHER_RACE = 'B18'
MULTI_RACIAL = 'B19'
HISPANIC = 'B21'
NOT_HISPANIC = 'B22'
WHITE_NOT_HISPANIC = 'B23'
CITIZENS_25_AND_OLDER = 'B24'
BACHELORS_OR_HIGHER = 'B25'
DETERMINED_POVERTY_STATUS = 'B26'
BELOW_POVERTY_LEVEL = 'B27'
HOUSEHOLDS = 'B28'
HUNDRED_K_HOUSEHOLDS = 'B29'

# urls found at: http://www.census.gov/data/tables/time-series/demo/voting-and-registration/electorate-profiles-2016.html
URL = 'http://www2.census.gov/programs-surveys/demo/tables/voting/{}.xlsx'

def write_object(f, name, o):
    f.write('export const {} = '.format(name))
    f.write(json.dumps(o, indent=2, sort_keys=True))
    f.write(';\n')

states = {}
for state, e_votes in E_VOTES.items():
    print(state)
    fn = os.path.join(DATA_DIR, state+'.xlsx')

    if not os.path.exists(fn):
        url = URL.format(state.replace(' ', ''))
        r = requests.get(url)
        with open(fn, 'wb') as f:
            f.write(r.content)

    wb = load_workbook(fn)
    ws = wb[wb.sheetnames[0]]

    states[state] = {
        'name': state,
        'elVotes': e_votes,
        'votingAgePop': ws[VOTING_AGE_POP].value,
    }

elVotes = sum([state['elVotes'] for state in states.values()])
votingAgePop = sum([state['votingAgePop'] for state in states.values()])
nation = {
    'name': 'United States',
    'elVotes': elVotes,
    'votingAgePop': votingAgePop,
}

for state in states.values():
    state['elVoteRatio'] = state['elVotes'] / nation['elVotes']
    state['elVotePercent'] = state['elVoteRatio'] * 100
    state['votes'] = state['elVoteRatio'] * nation['votingAgePop']
    state['votesPerPerson'] = state['votes'] / state['votingAgePop']

with open(JS_FILE, 'w') as f:
    write_object(f, 'states', states)
    write_object(f, 'nation', nation)
